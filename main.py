import requests
import json
import openpyxl
import time
import openpyxl.styles
import os

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0',
    'Authorization':'Bearer eyJhbGciOiJSUzI1NiIsImtpZCI6IkEyOTQ4RUY3Nzk3RkZFNkQ0OTcyOTQ0ODY5OTU3MkU5IiwidHlwIjoiYXQrand0In0.eyJuYmYiOjE3MTU1MDIzNTYsImV4cCI6MTcxNTUwNTk1NiwiaXNzIjoiaHR0cHM6Ly9vcGVuaWQuY2M5OC5vcmciLCJjbGllbnRfaWQiOiI5YTFmZDIwMC04Njg3LTQ0YjEtNGMyMC0wOGQ1MGE5NmU1Y2QiLCJzdWIiOiI3MzUzNzYiLCJhdXRoX3RpbWUiOjE3MTM1ODk4MzEsImlkcCI6ImxvY2FsIiwidW5pcXVlX25hbWUiOiJBS29uamFjXyIsIm5hbWUiOiJBS29uamFjXyIsImZvcnVtLnByaXZpbGVnZSI6NCwianRpIjoiMDFDODM4QzdDNzFGMTkzMTU3OEY1MzRFNjBBNEI3NzEiLCJpYXQiOjE3MTM1ODk4MzEsInNjb3BlIjpbImNjOTgtYXBpIiwib3BlbmlkIiwib2ZmbGluZV9hY2Nlc3MiXSwiYW1yIjpbIkNDOTgiXX0.E7TcnkC-FG2NWO1dULuQRhZyuSsk0EGTM8Q4uCbNE4KM0fggNXtmcN1HB0a_Kjfdr4i8yYofNuZUzWOKNQiagFNB253i1ixOdCDZdyCb6lWnocbotczwbZMyORt_YLox0K7wUD4AR67Y_U7rwQfP0Ceu7lPGHJM-jNbZL2yyxdE7m7y9wSPC5WMVP-7pCPTSOqUo9RXGONQNC9nS735BfBoEYefpoLV0YbfxFGA-EYbnAW_azbZcRYa3vKdi1ZIBWk4yQTwNzhKYkVsZeIGh7vCAXpGEyEhkpZWy78tZ3qnft9ZlVCWt0KBl3bp1r3mY6BtJxIGDHrm3LboEUKNDgw'
}# 可以改成自己浏览器的Authorization
tmp_path = './tmp.txt'  # workbook_num
num_path = './num.txt'  # topic_num
workbook_name = 'cc98_study_{}.xlsx'
web_name = r'https://api.cc98.org'
# web_name = r'http://www-cc98-org-s.webvpn.zju.edu.cn:8001/' # 可改成webvpn连接cc98


def get_time(t0, t1, t2): # 得到下载时间
    t = t2 - t1
    h = int(t / 3600)
    m = int((t - h * 3600) / 60)
    s = int((t - h * 3600 - m * 60))
    print("section time:" + str(h) + ":" + str(m) + ":" + str(s))
    t = t2 - t0
    h = int(t / 3600)
    m = int((t - h * 3600) / 60)
    s = int((t - h * 3600 - m * 60))
    print("total time:" + str(h) + ":" + str(m) + ":" + str(s))


def get_new_tmp(): # 新的excel文件编号
    with open(tmp_path, 'r') as f:
        tmp = f.read()
    num_tmp = int(tmp)
    with open(tmp_path, 'w') as f:
        f.write(str(num_tmp+1))
    return num_tmp+1


def work(i, sheet):
    Url = web_name+r'/Topic/{}?sf_request_type=fetch'.format(i) # 帖子网址
    resp = requests.get(Url, headers=headers)
    resp.encoding = 'utf-8'
    html = resp.text
    # 特殊情况处理，防止退出
    if html == "topic_not_exists":
        print(str(i) + ":not exist")
        return
    elif html == "topic_is_deleted":
        print(str(i) + ":deleted")
        return
    try:
        js_html = json.loads(html)
        cnt = int(js_html["replyCount"] / 10) # 帖子中的评论包的数量

        tmp = []
        for j in range(0, cnt + 1):
            url = web_name+r'/Topic/{}/post?from={}&size=10&sf_request_type=fetch'.format(
                i, j * 10)  # 帖子中的评论包的网址
            res = requests.get(url, headers=headers)
            res.encoding = 'utf-8'
            htmls = res.text
            js_htmls = json.loads(htmls)
            for k in range(0, 10):
                try:
                    tmp.append([str(js_htmls[k]["floor"]), js_htmls[k]["content"],'赞', str(js_htmls[k]["likeCount"]), '踩', str(js_htmls[k]["dislikeCount"]), '发布时间', str(js_htmls[k]["time"])]) # 相关信息加入到列表
                except IndexError:
                    continue

        print(str(i) + ":saved") # 提示信息
        sheet.append(['topic', str(i), js_html["title"]]) # 保存到excel
        row = sheet[sheet.max_row]
        for r in row:
            r.font = openpyxl.styles.Font(name="微软雅黑", size=12, bold=True, italic=False)
        sheet.append(['观看', str(js_html["hitCount"]), '收藏', str(js_html["favoriteCount"]), '发布时间', js_html["time"]])
        row = sheet[sheet.max_row]
        for r in row:
            r.font = openpyxl.styles.Font(name="微软雅黑", size=10, bold=False, italic=True)

        for T in tmp:
            try:
                sheet.append(T)
            except openpyxl.utils.exceptions.IllegalCharacterError: # 非法字符
                print("IllegalCharacter:" + str(T[0]))
                continue

    except json.decoder.JSONDecodeError: # 编码错误，可能有各种原因导致，一般而言可能是Authorization需要更新
        print(str(i) + ":Authorization Error")
        return


def spider(workbook_num, board_num):
    t0 = time.time()
    t1 = time.time()
    workbook = openpyxl.Workbook() # 创建excel
    sheet = workbook.active
    sheet.title = 'cc98'
    web_url = web_name+r'/board/{}?sf_request_type=fetch'.format(board_num) # 版面网址
    resp = requests.get(web_url, headers=headers)
    resp.encoding = 'utf-8'
    html = resp.text
    js_html = json.loads(html)

    num = int(js_html["topicCount"] / 20) # 版面主题数量/20=版面页数=版面主题包数
    with open(num_path, 'r') as f:
        now = f.read()
    now_num = int(now) # 当前到了版面那一页，防止中途退出后从头开始爬
    for n in range(now_num + 1, num + 1):
        web_Url = web_name+r'/board/{}/topic?from={}&size=20&sf_request_type=fetch'.format(board_num, n * 20) # 版面中主题帖网址
        resp = requests.get(web_Url, headers=headers)
        resp.encoding = 'utf-8'
        html = resp.text
        js_html = json.loads(html)
        try:
            for k in range(0, 20):
                try:
                    work(js_html[k]["id"], sheet) # 得到主题帖编号，调用work
                except IndexError:
                    continue
        except json.decoder.JSONDecodeError:
            print("topic list Authorization error")
            continue
        if n % 5 == 0:  # 每100个帖子保存一下
            workbook.save(workbook_name.format(workbook_num))
            get_time(t0, t1, time.time())
            t1 = time.time()
            print(n)
            with open(num_path, 'w') as f: # 更新num
                f.write(str(n))
        if n % 500 == 0:  # 每个n有20个帖子，所以10000个帖子换一个excel文档
            workbook.remove(sheet)
            workbook.create_sheet('cc98')
            sheet = workbook['cc98']
            workbook_num = get_new_tmp()

    workbook.save(workbook_name.format(workbook_num))


if __name__ == "__main__":
    if not os.path.exists(num_path):
        with open(num_path, 'w') as f:
            f.write(str(0))
    if not os.path.exists(tmp_path):
        with open(tmp_path, 'w') as f:
            f.write(str(0))
    with open(tmp_path, 'r') as f:
        workbook_num = f.read()
    board_num = 68 # 版面编号，可更改
    spider(workbook_num, board_num)
